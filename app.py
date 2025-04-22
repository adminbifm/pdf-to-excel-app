import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def extraer_datos(pdf_file):
    cuentas_objetivo = [
        ("349", "TOTAL ACTIVOS CORRIENTES"),
        ("389", "TOTAL ACTIVOS INTANGIBLES"),
        ("599", "TOTAL DEL PASIVO"),
        ("698", "TOTAL PATRIMONIO NETO"),
        ("701", "UTILIDAD DEL EJERCICIO"),
        ("6999", "TOTAL INGRESOS"),
        ("7991", "TOTAL COSTOS"),
        ("314", "Locales"),
        ("316", "Locales"),
        ("318", "Locales"),
    ]

    resultado = []
    pattern = re.compile(r"^(.*?)\s+(\d{3,4})\s+([\d.,-]+)$")

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                for line in text.split("\n"):
                    match = pattern.match(line.strip())
                    if match:
                        descripcion = match.group(1).strip()
                        codigo = match.group(2)
                        valor = float(match.group(3).replace(",", ""))
                        for cod_ref, desc_ref in cuentas_objetivo:
                            if codigo == cod_ref and desc_ref.lower() in descripcion.lower():
                                resultado.append([descripcion, codigo, valor])
                                break

    df = pd.DataFrame(resultado, columns=["Descripción", "Código", "Valor"])
    valor_cuentas = df[df["Código"].isin(["314", "316", "318"])]["Valor"].sum()
    df_cxc = pd.DataFrame([["CUENTAS POR COBRAR", "CXC", valor_cuentas]], columns=df.columns)
    ingresos = df[df["Código"] == "6999"]["Valor"].sum()
    costos = df[df["Código"] == "7991"]["Valor"].sum()
    df_gb = pd.DataFrame([["GANANCIA BRUTA", "GB", ingresos - costos]], columns=df.columns)

    return pd.concat([df, df_cxc, df_gb], ignore_index=True)

# App Streamlit
st.title("📄 Convertidor de Declaraciones PDF a Excel")

pdf_file = st.file_uploader("Sube tu declaración en PDF", type=["pdf"])

if pdf_file is not None:
    st.info("Procesando archivo...")

    df_final = extraer_datos(pdf_file)

    # Guardar hoja DATA-BRUTO
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_final.to_excel(writer, sheet_name="DATA-BRUTO", index=False)

    output.seek(0)
    wb = load_workbook(output)
    ws = wb["DATA-BRUTO"]

    # Obtener valores desde la hoja
    def get_val(codigo):
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[1]) == str(codigo):
                return float(row[2])
        return 0

    ingresos = get_val(6999)
    activos_corr = get_val(349)
    intangibles = get_val(389)
    pasivo = get_val(599)
    patrimonio = get_val(698)
    utilidad = get_val(701)
    ganancia = next((row[2] for row in ws.iter_rows(min_row=2, values_only=True) if row[1] == "GB"), 0)

    # Crear hoja Decisioning
    ws2 = wb.create_sheet("Decisioning")
    ws2["B1"] = "Knockout Rules"
    ws2["B1"].font = Font(bold=True, size=14)

    headers = ["Sr No", "Parameters", "Criteria", "Actual Values", "Pass/Fail"]
    ws2.append([""])  # Fila vacía para llegar a B3
    ws2.append([""] + headers)

    # Contenido
    contenido = [
        [1, "Minimum Annual revenue $5,000,000", ">=\$200,000", ingresos, '=SI(E4>=200000,"Pass","Fail")'],
        [2, "Negative bank balance days in the last 6 months", "<=5", "No se cuenta con la información", ""],
        [3, "Liquidity Runway", ">=6 Months", round(activos_corr / pasivo, 2) if pasivo != 0 else "", '=SI(E6>=6,"Pass","Fail")'],
        [4, "If Tangible Net Worth is negative, business must be profitable", "N/A", patrimonio - intangibles, '=SI(E7>=0,"Pass","Fail")'],
        [5, "Net Income Margin", "<-5%", round(utilidad / ingresos, 4) if ingresos != 0 else "", '=SI(E8>=-0.05,"Pass","Fail")'],
        [6, "Current liabilities must not exceed 60% of annual revenue", "<=60%", round(pasivo / ingresos, 4) if ingresos != 0 else "", '=SI(E9<=0.6,"Pass","Fail")'],
        [7, "Gross Margin", ">10%", round(ganancia / ingresos, 4) if ingresos != 0 else "", '=SI(E10>0.1,"Pass","Fail")'],
        [8, "Minimum time in Business (In Years)", ">=3 Years", "", '=SI(E11>=3,"Pass","Fail")'],
        [9, "Minimum Experian Intelliscore", "N/A", "", ""],
        [10, "Not delinquent on any Slope obligations or gone more than 15 days delinquent on any prior Slope obligations", "", "Aprobado Crédito", '=SI(E13="Aprobado Crédito","Pass","Fail")'],
    ]

    start_row = 4
    for i, row in enumerate(contenido):
        ws2.cell(row=start_row + i, column=2, value=row[0])
        ws2.cell(row=start_row + i, column=3, value=row[1])
        ws2.cell(row=start_row + i, column=4, value=row[2])
        ws2.cell(row=start_row + i, column=5, value=row[3])
        if isinstance(row[4], str) and row[4].startswith("=SI"):
            ws2.cell(row=start_row + i, column=6).value = row[4]
        else:
            ws2.cell(row=start_row + i, column=6, value=row[4])

    # Anchos
    ws2.column_dimensions["C"].width = 62
    ws2.column_dimensions["D"].width = 15.71
    ws2.column_dimensions["E"].width = 31.57

    # Estilo encabezado
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws2["B3":"F3"][0]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Alineación celdas restantes
    for row in ws2.iter_rows(min_row=4, min_col=2, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")

    # Guardar archivo final
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.success("✅ Archivo procesado con éxito")
    st.download_button(
        label="📥 Descargar Excel",
        data=final_output,
        file_name="declaracion_convertida.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
